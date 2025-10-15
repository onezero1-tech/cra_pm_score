# encoding:utf-8

from fastapi import FastAPI, UploadFile, Form
from fastapi.responses import StreamingResponse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
import io
import zipfile

app = FastAPI()

def get_df(data_content: bytes, sheet_name: str, usecols: list):
    data_io = io.BytesIO(data_content)
    df = pd.read_excel(data_io, sheet_name=sheet_name, usecols=usecols, header=0, engine='calamine')
    return df

def copy_style_no_fill(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.protection = copy(src_cell.protection)
    # 不复制 fill（颜色）

@app.get("/")
def read_root():
    return {"message": "Welcome to Excel Processor"}

@app.post("/process")
async def process_files(
    data_file: UploadFile,  # 必传：数据 Excel 文件
    template_file: UploadFile | None = None,  # 可选：模板 Excel 文件
    sheet_name: str = Form("02-项目汇总表"),  # 可选参数，默认值
    usecols: str = Form("D,E,F,I,K"),  # 可选，逗号分隔
    header_row: int = Form(1),  # 可选
    data_start: int = Form(4)   # 可选
):
    if not template_file:
        return {"error": "模板文件是必需的，请上传。"}

    usecols_list = [col.strip() for col in usecols.split(',')]

    data_content = await data_file.read()
    template_content = await template_file.read()

    df = get_df(data_content, sheet_name, usecols_list)

    template_io = io.BytesIO(template_content)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for k_value, sub_df in df.groupby(usecols_list[-1]):
            safe_name = str(k_value).replace('/', '_')
            out_io = io.BytesIO()

            # 为每个分组加载模板（重置 IO）
            template_io.seek(0)
            wb = load_workbook(template_io)
            ws_tpl = wb['A']  # 假设模板 sheet 是 'A'

            for d_value, mini_df in sub_df.groupby(usecols_list[0]):
                sheet_name_d = str(d_value)
                if sheet_name_d in wb.sheetnames:
                    wb.remove(wb[sheet_name_d])
                ws = wb.copy_worksheet(ws_tpl)
                ws.title = sheet_name_d

                # 写入数据（假设 E->col1, F->col2, I->col3，与原代码一致）
                # 如果 usecols 变化，可动态调整，但这里固定匹配原逻辑
                for r_idx, (_, row) in enumerate(mini_df.iterrows(), start=data_start):
                    e_cell = ws.cell(row=r_idx, column=1, value=row[usecols_list[1]])
                    f_cell = ws.cell(row=r_idx, column=2, value=row[usecols_list[2]])
                    i_cell = ws.cell(row=r_idx, column=3, value=row[usecols_list[3]])
                    tpl_row = data_start
                    copy_style_no_fill(ws_tpl.cell(row=tpl_row, column=1), e_cell)
                    copy_style_no_fill(ws_tpl.cell(row=tpl_row, column=2), f_cell)
                    copy_style_no_fill(ws_tpl.cell(row=tpl_row, column=3), i_cell)

            # 删除模板 sheet
            if 'A' in wb.sheetnames:
                wb.remove(wb['A'])

            wb.save(out_io)
            out_io.seek(0)
            zipf.writestr(f'{safe_name}.xlsx', out_io.getvalue())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=processed_excels.zip"}
    )