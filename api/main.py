# main.py # encoding:utf-8
import io
import logging
from copy import copy
from typing import List
import os

import pandas as pd
import py7zr
import zipfile
import re
from fastapi import FastAPI, File, Form, HTTPException,UploadFile
from fastapi.middleware.cors import CORSMiddleware # <--- 修正 1: 导入 CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)

app = FastAPI()

# CORS Middleware is still important
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# 在文件头部追加


# ---------- 新增：zip/7z 解压工具 ----------
def extract_archive(archive_bytes: bytes, file_name: str) -> dict[str, bytes]:
    """返回 文件名 → 文件内容 的字典（只保留 xls/xlsx）"""
    ret: dict[str, bytes] = {}
    lower_name = file_name.lower()
    if lower_name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as z:
            for info in z.infolist():
                if info.filename.lower().endswith((".xls", ".xlsx")):
                    ret[info.filename] = z.read(info)
    elif lower_name.endswith(".7z"):
        with py7zr.SevenZipFile(io.BytesIO(archive_bytes), mode="r") as z:
            for fname, bio in z.readall().items():
                if fname.lower().endswith((".xls", ".xlsx")):
                    ret[fname] = bio.read()
    else:
        raise ValueError("只支持 .zip / .7z 压缩包")
    if not ret:
        raise ValueError("压缩包内未找到 Excel 文件")
    return ret

# ---------- 新增：/merge 接口 ----------
@app.post("/merge")
async def merge_archive(archive_file: UploadFile = File(...)):
    try:
        archive_bytes = await archive_file.read()
        file_map = extract_archive(archive_bytes, archive_file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解压失败: {e}")

    all_frames: list[pd.DataFrame] = []
    for file_name, file_bytes in file_map.items():
        # 取文件名 key（按“-”分割第 5 段）
        index_pos = file_name.rfind("/")
        base = file_name
        if index_pos != -1:
            base = file_name[index_pos + 1:]
        file_key = base.split("-")[4] if len(base.split("-")) > 4 else base
        if file_key.endswith(".xlsx"):
            file_key = file_key[:-5]
        logging.error(f"file name is {file_name}, file key is {file_key}")

        try:
            xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="calamine")
            hs_sheets = [s for s in xl.sheet_names if str(s).startswith("HS")]
            for sheet in hs_sheets:
                df = pd.read_excel(xl, sheet_name=sheet, header=None)
                if df.shape[1] <= 1:
                    continue
                df = df.dropna(axis=1, how="all").iloc[3:, 1:]   # 删掉空列，再去掉前 3 行、第 1 列
                df = df[df.iloc[:, 0].notna()]                   # 第一列非空
                df.insert(0, "Sheet", sheet)
                df.insert(0, "File", file_key)
                all_frames.append(df)
        except Exception as e:
            # 单个文件出错继续跑
            logging.warning("读取 %s 失败: %s", file_name, e)

    if not all_frames:
        raise HTTPException(status_code=400, detail="未找到任何符合条件的 Sheet")

    final = pd.concat(all_frames, ignore_index=True)
    out_io = io.BytesIO()
    final.to_excel(out_io, index=False)
    out_io.seek(0)
    return StreamingResponse(
        out_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=merged.xlsx"}
    )

def parse_numeric_positions(usecols_str: str) -> List[int]:
    """
    将 "1,4,5" 解析为 0 起始的整数列表 [0,3,4]
    抛出 ValueError 如果包含非正整数
    """
    parts = [p.strip() for p in usecols_str.split(',') if p.strip() != '']
    if not parts:
        return []
    positions = []
    for p in parts:
        if not p.isdigit():
            raise ValueError(f"非法列索引: {p}. 请使用逗号分隔的正整数，例如 '1,4,5'.")
        n = int(p)
        if n < 1:
            raise ValueError(f"列索引必须 >= 1: {p}")
        positions.append(n - 1)  # 转为 0-based
    return positions

def get_df_by_position_stream(data_content: bytes, sheet_name: str, positions: List[int], header: int = 0):
    """
    读取整个 sheet，然后按列位置选择所需列返回 DataFrame。
    为降低内存，尽量只保留必要列。 pandas.read_excel 不支持只按位置读取，
    因此读取时先加载全部列头行，然后按 usecols 字母再读取数据是一种减少内存的方法。
    这里为通用性选择直接读取全部并切片，但随后释放不必要引用以降低内存驻留。
    """
    bio = io.BytesIO(data_content)
    # header 参数：传入的是 pandas.read_excel 的 header（0-based 行索引或 None）
    df_all = pd.read_excel(bio, sheet_name=sheet_name, header=header, engine='calamine')
    max_idx = df_all.shape[1] - 1
    for pos in positions:
        if pos > max_idx:
            raise IndexError(f"请求的列索引 {pos+1} 超出表格列数 {max_idx+1}")
    selected = df_all.iloc[:, positions].copy()
    # 释放大 DataFrame 引用（尽量降低内存）
    del df_all
    return selected

def copy_style_no_fill(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.protection = copy(src_cell.protection)
    # 不复制 fill（颜色）

@app.get("/health")
def read_root():
    return {"message": "Welcome to Excel Processor"}

@app.post("/process")
async def process(
    data_file: UploadFile = File(...),
    template_file: UploadFile = File(...),
    sheet_name: str = Form("02-项目汇总表"),
    usecols: str = Form("4,5,6,9,11"),
    header_row: int = Form(1),
    data_start: int = Form(4),
):
    try:
        positions = parse_numeric_positions(usecols)
    except ValueError as e:
        logging.error(f"Invalid positions: {e}")
        # <--- 修正 2: 确保 HTTPException 被导入
        raise HTTPException(status_code=400, detail=str(e))

    data_content = await data_file.read()
    template_content = await template_file.read()

    try:
        df = get_df_by_position_stream(data_content, sheet_name, positions, header=header_row - 1)
    except Exception as e:
        logging.error(f"Error reading Excel data: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"读取数据文件失败: {e}")

    logging.info("df shape is {}".format(df.shape))
    if df.empty:
        raise HTTPException(status_code=400, detail="未从指定列和工作表中提取到任何数据。")

    # 按提取的最后一列进行分组 (例如 usecols="4,5,6,9,11" 中的第11列)
    # <--- 修正 3: 更新注释使其与代码行为一致
    group_col = df.columns[-1]

    seven_zip_buffer = io.BytesIO()
    with py7zr.SevenZipFile(seven_zip_buffer, mode='w') as archive:
        for k_value, sub_df in df.groupby(group_col, sort=False):
            safe_name = str(k_value).replace('/', '_')
            out_io = io.BytesIO()
            tpl_io = io.BytesIO(template_content)
            wb = load_workbook(tpl_io)
            ws_tpl = wb['A'] if 'A' in wb.sheetnames else wb[wb.sheetnames[0]]

            first_col = df.columns[0]
            for d_value, mini_df in sub_df.groupby(first_col, sort=False):
                sheet_name_d = str(d_value)
                if len(sheet_name_d) > 31: # Excel sheet name limit
                    sheet_name_d = sheet_name_d[:31]
                
                if sheet_name_d in wb.sheetnames:
                    wb.remove(wb[sheet_name_d])
                ws = wb.copy_worksheet(ws_tpl)
                ws.title = sheet_name_d

                for r_idx, (_, row) in enumerate(mini_df.iterrows(), start=data_start):
                    v1 = row.iloc[1] if len(row) > 1 else None
                    v2 = row.iloc[2] if len(row) > 2 else None
                    v3 = row.iloc[3] if len(row) > 3 else None
                    if v1 is not None:
                        c = ws.cell(row=r_idx, column=1, value=v1)
                        copy_style_no_fill(ws_tpl.cell(row=data_start, column=1), c)
                    if v2 is not None:
                        c = ws.cell(row=r_idx, column=2, value=v2)
                        copy_style_no_fill(ws_tpl.cell(row=data_start, column=2), c)
                    if v3 is not None:
                        c = ws.cell(row=r_idx, column=3, value=v3)
                        copy_style_no_fill(ws_tpl.cell(row=data_start, column=3), c)

            if ws_tpl.title in wb.sheetnames:
                try:
                    wb.remove(ws_tpl)
                except Exception:
                    pass
            wb.save(out_io)
            out_io.seek(0)
            archive.writestr(out_io.read(), str(safe_name) + '.xlsx')

    seven_zip_buffer.seek(0)
    logging.info("Processing finished. Sending response.")
    return StreamingResponse(
        seven_zip_buffer,
        media_type="application/x-7z-compressed",
        headers={"Content-Disposition": "attachment; filename=processed_excels.7z"}
    )